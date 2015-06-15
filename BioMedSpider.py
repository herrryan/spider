#-*- coding: UTF-8 -*-

import sys
import time
import urllib
import urllib2
import requests
import re
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook

reload(sys)
sys.setdefaultencoding('utf8')


def biomed_spider():
    page_num=0;
    count=1;
    project_lists=[];
    try_times=0;
    #Some User Agents
    headers=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
         {'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
         {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}];

    #while(1):
    url="http://www.biomed.ee.ethz.ch/education/student_projects/biomechanics/cartilage";
    print "Pulling information from [%s]" % url;
    sleep_time = np.random.rand()*2;
    print "Sleeping for %f" % sleep_time;
    time.sleep(sleep_time);
    
    #Last Version
    try:
        req = urllib2.Request(url, headers=headers[1]);
        source_code = urllib2.urlopen(req).read();
        plain_text=str(source_code);
    except (urllib2.HTTPError, urllib2.URLError), e:
        print e
        #continue
    soup = BeautifulSoup(plain_text);
    src_project_list = soup.find('table', {'class': 'silvatable list'});
    if src_project_list == None:
   		print "There is no project for this Lab yet.";
   		return;
    
   	
    count = 0;
    for project_info in src_project_list.findAll('td'):
        #print "Does it contain <a> tag: %s" % project_info.find('a');
        #print "project_info: %s" % project_info;
        if project_info.find('a') != None:
            detail_info = project_info.findAll('a');
            project_name = detail_info[0].get_text().replace('\n', ' ');
            project_supervisor = detail_info[1].get_text().replace('\n', ' ');
            project_pdf = detail_info[0].get('href');
            project_supervisor_link = detail_info[1].get('href');
            count += 1;
            print "Project number %d" % count;
            print "project name: %s" % project_name;
            print "project supervisoer: %s" % project_supervisor;
            print "project pdf link %s" % project_pdf;
            print "project Supervisor link %s" % project_supervisor_link;
            project_lists.append([project_name, project_supervisor, project_pdf, project_supervisor_link]);
    return project_lists

def print_project_lists_excel(project_lists):
    wb=Workbook(optimized_write=True);
    ws=[];
    ws.append(wb.create_sheet(title="Available Project Information"));
    ws[0].append(['Number', 'Project Name', 'Project Supervisor', 'Project PDF', 'Project Supervisor Link']);
    count=1;
    for project in project_lists:
        ws[0].append([count,project[0],project[1], project[2], project[3]]);
        count+=1;
    save_path='Available_project_information.xlsx';
    wb.save(save_path)



if __name__=='__main__':
    project_lists = biomed_spider();
    print_project_lists_excel(project_lists);